import requests 
from openpyxl import Workbook
from bs4 import BeautifulSoup
from konlpy.tag import Kkma
from wordcloud import WordCloud, ImageColorGenerator
import matplotlib.pyplot as plt
import numpy as np
from PIL import Image


kkma = Kkma()

def get_soup(url) :
	headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36'}
	res = requests.get(url, headers=headers)
	res.raise_for_status()
	soup = BeautifulSoup(res.text, 'lxml')
	return soup

def get_epilogue() :
	epil_detail_url = 'https://www.autoplus.co.kr/story/WUCO010002.rb?boardNo='
	
	# write excel
	write_wb = Workbook()
	write_ws = write_wb.active
	write_ws.append(['게시물 번호', '평점', '제목', '내용', 'URL(링크)'])
	
	detail_cont_list = []

	# col
	col_idx = 1
	for num in range(1, 100) : 
		soup = get_soup(epil_detail_url + str(num))	
		# 해당 번호의 게시물 존재하는지 확인
		try : 
			title = soup.find('div', attrs={'class': 'subject'}).find('p').get_text().strip()
			detail_conts = soup.find('div', attrs={'class': 'cont-box'}).find_all('p')
			result_cont = ''
			for detail_cont in detail_conts :
				result_cont += detail_cont.get_text().strip()
			grade = soup.find('div', attrs={'class': 'grade-view'}).find('span', attrs={'class', 'num'}).get_text().strip()

			# excel
			for row_idx in range(1, 6) :
				if row_idx == 1 :
					write_ws.cell(col_idx, row_idx, num)
				elif row_idx == 2 :
					write_ws.cell(col_idx, row_idx, grade)
				elif row_idx == 3 : 
					write_ws.cell(col_idx, row_idx, title)
				elif row_idx == 4 :
					write_ws.cell(col_idx, row_idx, result_cont)
				elif row_idx == 5 :
					write_ws.cell(col_idx, row_idx, epil_detail_url+str(num))
			col_idx = col_idx + 1

			detail_cont_list.append(result_cont)

			# print(f'{num}')
			# print(f'{title} / 평점 : {grade}')
			# print(f'{result_cont}')
		except AttributeError as e : 
			pass
	

	
	# save excel file
	write_wb.save('C:/Users/PC/Documents/simbyungki/git/autoplus/autoplus_epilogue_analysis/epilogue.xlsx')
	return detail_cont_list

def read_img_from_url_and_return_matrix(url):
    response = requests.get(url)
    #print("binary file sample: {}".format(response.content[:20]))

    from PIL import Image 
    from io import BytesIO 

    img = Image.open(BytesIO(response.content))
    img_matrix = np.array(img)
    # plt.imshow(img_matrix)
    # img.save('aaa.png')
    return img_matrix

def make_ImageColoredWordcloud(text, img_matrix, outputfile_name):
    wc = WordCloud(background_color='white', max_words=300, mask=img_matrix, max_font_size=150, width=600, height=400, font_path='C:/Users/PC/Documents/simbyungki/git/autoplus/autoplus_epilogue_analysis/IropkeBatangM.woff')
    # wc = WordCloud(background_color='white', max_words=300, mask=img_matrix, max_font_size=110, random_state=42, width=600, height=400, font_path='C:/Users/PC/Documents/simbyungki/git/autoplus/autoplus_epilogue_analysis/IropkeBatangM.woff')
    # generate word cloud
    wc.generate(text)
    f = plt.figure(figsize=(20, 20))
    plt.imshow(wc.recolor(color_func=ImageColorGenerator(img_matrix)), interpolation='bilinear')
    plt.axis('off')
    plt.savefig(outputfile_name)

def get_nouns() :
	epilogue_list = get_epilogue()
	epilogue_noun_list = []
	result_text_list = ''
	result_cnt = {}
	except_keyword_list = ['오토플러스', '리본카', '리본', '오토']


	# make list
	# for epilogue in epilogue_list :
	# 	for keyword in kkma.nouns(epilogue) :
	# 		if len(keyword) > 1 :
	# 			epilogue_noun_list.append(keyword)


	# make text
	for epilogue in epilogue_list :
		for keyword in kkma.nouns(epilogue) :
			if (keyword not in except_keyword_list) and len(keyword) > 1 :
				result_text_list += ' '+keyword


	for noun in epilogue_noun_list :				
		try : 
			result_cnt[noun] += 1
		except : 
			result_cnt[noun] = 1
	
	# sort
	# re_sort = sorted(result_cnt.items(), key=(lambda x: x[1]), reverse=True)
	print(result_text_list)

	# wordcloud = WordCloud(
	# 	max_font_size=110, \
	# 	font_path='C:/Users/PC/Documents/simbyungki/git/autoplus/autoplus_epilogue_analysis/IropkeBatangM.woff', \
	# 	width=600, \
	# 	height=400, \
	# 	max_words=150, \
	# 	background_color='white' \
	# ).generate(result_text_list)

	# flg = plt.figure()
	# plt.imshow(wordcloud, interpolation='bilinear')
	# plt.axis('off')
	# plt.savefig('C:/Users/PC/Documents/simbyungki/git/autoplus/autoplus_epilogue_analysis/wordcloud.png')

	url = 'C:/Users/PC/Documents/simbyungki/git/autoplus/autoplus_epilogue_analysis/bg.png'
	img_matrix = read_img_from_url_and_return_matrix(url)
	make_ImageColoredWordcloud(result_text_list, img_matrix, 'C:/Users/PC/Documents/simbyungki/git/autoplus/autoplus_epilogue_analysis/coloredWordCloud.png')



if __name__ == '__main__' :
	get_nouns()